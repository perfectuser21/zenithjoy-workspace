# app/routers/admin.py
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Form
from sqlalchemy.orm import Session
from datetime import timedelta, datetime
import secrets
import csv
from io import BytesIO, StringIO
from typing import Any, List, Dict, Optional

import openpyxl
from ..db import get_db
from .. import models
from ..security import require_admin
from ..schemas import UserCreate, UserOut, InviteCreate, InviteOut, TextQuestionCreate, QuestionOut, TextImportOut, ImportErrorItem
from pydantic import BaseModel

router = APIRouter(prefix="/admin", tags=["admin"])

def now_utc():
    return datetime.utcnow()

@router.get("/status")
async def admin_status():
    return {"status": "admin_ok", "message": "管理API运行正常"}

@router.post("/users", response_model=UserOut, dependencies=[Depends(require_admin)])
def create_user(payload: UserCreate, db: Session = Depends(get_db)):
    existing = db.query(models.User).filter(models.User.phone == payload.phone).first()
    if existing:
        # Update name if needed, keep status
        existing.name = payload.name
        db.add(existing)
        db.commit()
        db.refresh(existing)
        return UserOut(id=existing.id, name=existing.name, phone=existing.phone, status=existing.status)
    user = models.User(name=payload.name, phone=payload.phone, status="PENDING")
    db.add(user)
    db.commit()
    db.refresh(user)
    return UserOut(id=user.id, name=user.name, phone=user.phone, status=user.status)

@router.post("/users/invite", response_model=InviteOut, dependencies=[Depends(require_admin)])
def create_invite(payload: InviteCreate, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.phone == payload.phone).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found; create user first")
    if user.status == "DISABLED":
        raise HTTPException(status_code=403, detail="User disabled")
    code = secrets.token_urlsafe(8).replace("-", "").replace("_", "")[:10].upper()
    expires_at = now_utc() + timedelta(days=7)
    inv = models.Invite(user_id=user.id, code=code, expires_at=expires_at)
    db.add(inv)
    db.commit()
    return InviteOut(phone=user.phone, code=code, expires_at=expires_at.isoformat())

@router.post("/questions/text", response_model=QuestionOut, dependencies=[Depends(require_admin)])
def create_text_question(payload: TextQuestionCreate, db: Session = Depends(get_db)):
    if payload.correct_label not in payload.labels:
        raise HTTPException(status_code=400, detail="correct_label must be in labels")

    q_payload = {
        "type": "text_classification",
        "text": payload.text,
        "labels": payload.labels,
        "correct_label": payload.correct_label,
        "explanation": payload.explanation or "",
    }
    q = models.Question(type="text_classification", stem=payload.stem, payload=q_payload, is_active=True)
    db.add(q)
    db.commit()
    db.refresh(q)
    return QuestionOut(id=q.id, type=q.type, stem=q.stem, payload=q.payload)


def _split_labels(raw: str) -> list[str]:
    raw = str(raw or "").strip()
    if not raw:
        return []
    for sep in ["|", ",", "，", ";", "；"]:
        if sep in raw:
            parts = [p.strip() for p in raw.split(sep)]
            return [p for p in parts if p]
    return [raw]

def _load_rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = content.decode("utf-8", errors="ignore")
    f = StringIO(text)
    reader = csv.DictReader(f)
    return [dict(r) for r in reader]

def _load_rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        if r is None:
            continue
        d: dict[str, Any] = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            if v is not None and str(v).strip() != "":
                empty = False
            d[h] = v
        if not empty:
            out.append(d)
    return out

@router.post("/questions/text/import", response_model=TextImportOut, dependencies=[Depends(require_admin)])
def import_text_questions(
    file: UploadFile = File(...),
    strict: bool = Query(default=True, description="If true, any error will cancel the whole import."),
    db: Session = Depends(get_db),
):
    """Bulk import text classification questions from CSV or XLSX."""
    filename = (file.filename or "").lower()
    content = file.file.read()
    if not content:
        return TextImportOut(ok=False, imported_count=0, failed_count=1, errors=[ImportErrorItem(row=0, message="Empty file")])

    if filename.endswith(".csv"):
        rows = _load_rows_from_csv(content)
    elif filename.endswith(".xlsx"):
        rows = _load_rows_from_xlsx(content)
    else:
        return TextImportOut(ok=False, imported_count=0, failed_count=1, errors=[ImportErrorItem(row=0, message="Unsupported file type. Use .csv or .xlsx")])

    required = {"stem", "text", "labels", "correct_label"}
    errors: list[ImportErrorItem] = []
    questions: list[models.Question] = []

    for row_idx, r in enumerate(rows, start=2):  # row 1 is header
        missing = [k for k in required if k not in r or r.get(k) in (None, "")]
        if missing:
            errors.append(ImportErrorItem(row=row_idx, message=f"Missing required fields: {', '.join(missing)}"))
            continue

        stem = str(r.get("stem")).strip()
        text_val = str(r.get("text")).strip()
        labels_raw = r.get("labels")
        correct_label = str(r.get("correct_label")).strip()
        explanation = str(r.get("explanation") or "").strip()

        labels = _split_labels(labels_raw)
        if not labels:
            errors.append(ImportErrorItem(row=row_idx, message="labels is empty"))
            continue
        if correct_label not in labels:
            errors.append(ImportErrorItem(row=row_idx, message="correct_label must be one of labels"))
            continue

        q_payload = {
            "type": "text_classification",
            "text": text_val,
            "labels": labels,
            "correct_label": correct_label,
            "explanation": explanation,
        }
        questions.append(models.Question(type="text_classification", stem=stem, payload=q_payload, is_active=True))

    if errors and strict:
        return TextImportOut(ok=False, imported_count=0, failed_count=len(errors), errors=errors)

    if questions:
        db.add_all(questions)
        db.commit()

    return TextImportOut(ok=(len(errors) == 0), imported_count=len(questions), failed_count=len(errors), errors=errors)

# =========================
# Python 题：创建 + 批量导入
# =========================

class AdminCreatePythonQuestion(BaseModel):
    stem: str
    starter_code: str
    expected_stdout: str
    strict: bool = False
    time_limit_sec: int = 2
    submodule: Optional[str] = None


@router.post("/questions/python")
def admin_create_python_question(
    payload: AdminCreatePythonQuestion,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """
    创建 Python 题（A类：输出 stdout 判分）
    """
    q = models.Question(
        type="python",
        stem=payload.stem,
        payload={
            "starter_code": payload.starter_code,
            "expected_stdout": payload.expected_stdout,
            "strict": payload.strict,
            "time_limit_sec": payload.time_limit_sec,
            "submodule": payload.submodule,
        },
    )
    db.add(q)
    db.commit()
    db.refresh(q)
    return {"id": q.id, "type": q.type, "stem": q.stem, "payload": q.payload}


@router.post("/questions/python/import")
def admin_import_python_questions(
    strict: bool = Form(True),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """
    批量导入 Python 题：CSV 或 XLSX
    必须列：
      - stem
      - starter_code
      - expected_stdout
    可选列：
      - strict
      - time_limit_sec
      - submodule
    """
    filename = (file.filename or "").lower()
    content = file.file.read()

    rows: List[Dict[str, Any]] = []

    # --- 读 CSV ---
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(StringIO(text))
        for r in reader:
            rows.append(dict(r))

    # --- 读 XLSX ---
    elif filename.endswith(".xlsx"):
        if openpyxl is None:
            raise HTTPException(status_code=500, detail="openpyxl not installed; cannot import xlsx")
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        for i in range(2, ws.max_row + 1):
            values = [ws.cell(row=i, column=j + 1).value for j in range(len(headers))]
            r = {headers[j]: values[j] for j in range(len(headers))}
            rows.append(r)
    else:
        raise HTTPException(status_code=400, detail="Only CSV or XLSX supported")

    def _to_bool(v):
        if v is None:
            return False
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        return s in ("1", "true", "yes", "y")

    errors = []
    created_ids = []

    for idx, r in enumerate(rows, start=2):
        stem = (r.get("stem") or "").strip()
        starter_code = r.get("starter_code") or r.get("code") or ""
        expected_stdout = r.get("expected_stdout") or r.get("expected") or ""

        if not stem or not str(starter_code).strip() or expected_stdout is None:
            errors.append({"row": idx, "error": "stem/starter_code/expected_stdout required"})
            continue

        q_strict = _to_bool(r.get("strict")) if r.get("strict") is not None else False

        tls = r.get("time_limit_sec")
        try:
            time_limit_sec = int(tls) if tls is not None and str(tls).strip() != "" else 2
        except Exception:
            time_limit_sec = 2

        submodule = r.get("submodule")
        if submodule is not None:
            submodule = str(submodule).strip() or None

        q = models.Question(
            type="python",
            stem=stem,
            payload={
                "starter_code": str(starter_code),
                "expected_stdout": str(expected_stdout),
                "strict": q_strict,
                "time_limit_sec": time_limit_sec,
                "submodule": submodule,
            },
        )
        db.add(q)
        db.flush()  # 先拿到id
        created_ids.append(q.id)

    # strict=True：有错就整批回滚
    if errors and strict:
        db.rollback()
        return {"ok": False, "strict": True, "errors": errors, "created": 0}

    db.commit()
    return {"ok": True, "strict": strict, "errors": errors, "created": len(created_ids), "ids": created_ids}